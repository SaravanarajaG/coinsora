from flask import Flask, render_template
import openpyxl

app = Flask(__name__)

def load_items_by_category():
    wb = openpyxl.load_workbook("storage.xlsx")
    categories = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        items = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            # ✅ Create unique ID using sheet name
            unique_id = f"{sheet_name}_{row[0]}"

            item = {
                "id": unique_id,
                "title": row[1],
                "author": row[2],
                "price": row[3],
                "image": row[4] if len(row) > 4 and row[4] else "",
                "category": sheet_name,
                "image2": row[6] if len(row) > 6 and row[6] else "",
                "image3": row[7] if len(row) > 7 and row[7] else "",
                "description": row[8] if len(row) > 8 and row[8] else "",
                "image4": row[9] if len(row) > 9 and row[9] else "",
                "image5": row[10] if len(row) > 10 and row[10] else "",
            }

            if not row[0] or not row[1]:
                continue

            items.append(item)

        categories[sheet_name] = items

    return categories


@app.route("/")
def home():
    categories = load_items_by_category()
    return render_template("index.html", categories=categories)


@app.route("/category/<category_name>")
def category_page(category_name):
    categories = load_items_by_category()
    items = categories.get(category_name)
    if not items:
        return "No items found for this category", 404
    return render_template("category.html", category=category_name, items=items)


# ✅ Make item_id a STRING instead of int
@app.route("/item/<item_id>")
def item_detail(item_id):
    categories = load_items_by_category()

    for cat_items in categories.values():
        for item in cat_items:
            if item["id"] == item_id:
                return render_template("details.html", item=item)

    return "Item not found", 404


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
